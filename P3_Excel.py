# [IS303 P3] Dakota Dowd

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font

fileName = input("Input the name of the file (remember to include .xlsx): ")
myWorkbook = load_workbook(fileName)

currSheet = myWorkbook.active

currSheet.title = "Algebra"

courseList = []

iRow = 2
iCol = "A"
lastCourse = ""

for row in currSheet.iter_cols(min_row=2, max_col=1, values_only=True):
    for currCourse in row:

        if currCourse != lastCourse:
            # append to list
            courseList.append(currCourse)
        
        lastCourse = currCourse

print(courseList)
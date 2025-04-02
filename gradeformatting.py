# Lincoln Adams, Dakota Dowd, Caleb Caten, Isaac Pratte, and Josh Knight
# IS 303 Section 004
# P3 Formatting Grades in Excel
# Description: Program will automatically format and organize groups of student data.

import openpyxl # Importing the needed modules to work with Excel in Python
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


formattedGrades = Workbook() # Creating a new Workbook
formattedGrades.remove(formattedGrades.active) # Deleting the default sheet

sFileName = input(f"\nPlease enter the file name of the data you would like to organize, followed by (.xlsx): ") # Multiple sheets can be organized instead of hardcoding one
dataWorkbook = openpyxl.load_workbook(filename = sFileName) # Loading the disorganized data into a new workbook
dataWorksheet = dataWorkbook.active # Creating a sheet object with the disorganized data

class Student(): # Creating student class

    def __init__(self, first, last, class_name, id, grade): # Constructor takes values for first and last name, class name, id number and grade
        self.first = first
        self.last = last
        self.id = id
        self.class_name = class_name
        self.grade = grade

lstStudents = [] # List of students to save the objects

sClassName = "" # Initializing class name value
for iRow in range(2, dataWorksheet.max_row + 1): # Goes through all rows of data in a sheet and saves only unique class names
    if ( dataWorksheet.cell(row=iRow,column=1).value != sClassName):
        sClassName = dataWorksheet.cell(row=iRow,column=1).value
        formattedGrades.create_sheet(sClassName) # Creates a new sheet for every unique class name

bold_font = Font(bold=True) # Bold font variable

for iSheetIndex in range(len(formattedGrades.sheetnames)): # Loops through each sheet
    currWS = formattedGrades.worksheets[iSheetIndex]
    currWS.append(["Last Name", "First Name", "Student ID", "Grade", None, "Summary Statistics", "Value"]) # Adds the headings in the first row

    for iColWidth in range(1, currWS.max_column + 1): # Adjusts the column width and bolds the headings as long as there is a value in the first row
        cell_value = currWS.cell(row=1, column=iColWidth).value
        if cell_value is not None:
            currWS.column_dimensions[get_column_letter(iColWidth)].width = len(str(cell_value)) + 5
            currWS.cell(row = 1, column = iColWidth).font = bold_font
    
iRow = 0
for iRow in range(2, dataWorksheet.max_row + 1):
    
    # Loops through each row of each sheet and saves the student information to a corresponding variable
    sClassName = dataWorksheet.cell(row = iRow, column = 1).value
    sLastName, sFirstName, sStudID = dataWorksheet.cell(row = iRow, column = 2).value.split("_")
    iGrade = dataWorksheet.cell(row = iRow, column = 3).value

    if sLastName and sFirstName and sStudID and iGrade: # Only creating and appending the new object if there are values for these variables

        oStudent = Student(sFirstName, sLastName, sClassName, sStudID, iGrade) # Stores those variables to attributes in the student object
        lstStudents.append(oStudent) # Saves each student object in a list

for iStud in range(len(lstStudents)):
    currWS = formattedGrades[lstStudents[iStud].class_name]

    # For each student object in a given class, appends the data to the next available row
    currWS.append([lstStudents[iStud].last, lstStudents[iStud].first, lstStudents[iStud].id, lstStudents[iStud].grade]) 

iSheetIndex = 0
for iSheetIndex in range(len(formattedGrades.sheetnames)):
    currWS = formattedGrades.worksheets[iSheetIndex]

    # Writes the summary statistics categories in each sheet
    currWS["F2"] = "Highest Grade" 
    currWS["F3"] = "Lowest Grade"
    currWS["F4"] = "Mean Grade"
    currWS["F5"] = "Median Grade"
    currWS["F6"] = "Number of Students"

    # Calculates the summary statistics for each 
    currWS["G2"] = f'=MAX(D2:D{currWS.max_row})'
    currWS["G3"] = f'=MIN(D2:D{currWS.max_row})'
    currWS["G4"] = f'=AVERAGE(D2:D{currWS.max_row})'
    currWS["G5"] = f'=MEDIAN(D2:D{currWS.max_row})'
    currWS["G6"] = f'=COUNTA(D2:D{currWS.max_row})'

    # Adds an auto filter to each row
    currWS.auto_filter.ref = f"A1:D{currWS.max_row}"

sNewFileName = input(f"\nPlease enter the name of the new file you would like to save, followed by (.xlsx): ") # Saves file as user created file name
formattedGrades.save(filename = sNewFileName) # Saves
formattedGrades.close() # Closes
